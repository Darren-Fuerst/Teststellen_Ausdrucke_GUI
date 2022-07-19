import subprocess
import sys
import os
from typing import Tuple
import pandas as pd
import json
import csv
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from modules.html_template import *

def __style_excel(df, name):
    wb = openpyxl.load_workbook(name)
    ws1 = wb.active

    #add border
    # needs to be added first so we can change special values borders later on
    thin = Side(border_style="thin", color="000000")
    for cols in ws1.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None):
        for cell in cols:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    #Duplikate markieren mit first und last keyword zu Series.duplicated()
    ws1 = __mark_duplicates(df, ws1, "first")
    ws1 = __mark_duplicates(df, ws1, "last")
    
    #Unmögliche Emails markieren
    ws1 = __mark_email(df["Email"], ws1)

    

    wb.save(filename = name)


def __mark_email(mail_series, worksheet):
    for i in range(len(mail_series)):
        if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", mail_series.iloc[i]):
            # i + 2 because excel is index from 1 and it also has headers so when Indexing from 0 like in the Dataframe
            # it is actually 2 indices above in the excel sheet
            # Column 10 is the mail column
            worksheet = __mark_cell(worksheet, i + 2, 10)
    return worksheet

def __mark_duplicates(df,worksheet, keep):
    duplicates = (df['Vorname'] + df["Familienname"]).duplicated(keep=keep)
    for i in range(1, len(duplicates) + 1):
        if duplicates[i]:
            worksheet = __mark_cell(worksheet, i+1, 1)
    return worksheet

def __mark_cell(worksheet, x, y):
    thick = Side(border_style="thick", color="000000")
    worksheet.cell(row= x, column=y).font = Font(color="FF0400")
    worksheet.cell(row= x, column=y).border = Border(top=thick, left=thick, right=thick, bottom=thick)
    return worksheet



def clean_export(path):
    """
    Cleans up the export so it is able to be read in by pandas.read_csv() 
    Sometimes the export has more columns than it should have because the csv provider doesn't escape "special signs"

    path: the Path to the csv to be cleaned
    """
    with open(path, "r", encoding="latin1") as file:
        file = csv.reader(file, delimiter=";")
        csv_list = []

        #grab headers 
        headers = next(file)

        for row in file:
            #cut off anything beyond the last header
            csv_list.append(row[0:len(headers)])
        with open(path, "w", encoding="latin1") as f:
            write = csv.writer(f, delimiter=";", lineterminator="\n")
            write.writerow(headers)
            write.writerows(csv_list)

def create_needed_folders():
    folders = [r"./Altdorf", r"./Hersbruck", r"./Lauf",r"./Altdorf/Listen", r"./Hersbruck/Listen", r"./Lauf/Listen", r"./Logs", r"./Export"]
    for folder in folders:
        os.makedirs(folder, exist_ok=True)


def open_files(Teststelle):
    if Teststelle == "Altdorf":
        files = ["./Altdorf/ALD_Schnelltests.html","./Altdorf/ALD_PCRs.html","./Altdorf/Listen/ALD_Liste_PCR.xlsx","./Altdorf/Listen/ALD_Liste_Schnell.xlsx"]
    elif Teststelle == "Lauf":
        files = ["./Lauf/LAU_Schnelltests.html","./Lauf/LAU_PCRs.html","./Lauf/Listen/LAU_Liste_PCR.xlsx","./Lauf/Listen/LAU_Liste_Schnell.xlsx"]
    elif Teststelle == "Hersbruck":
        files = ["./Hersbruck/HEB_Schnelltests.html","./Hersbruck/HEB_PCRs.html","./Hersbruck/Listen/HEB_Liste_PCR.xlsx","./Hersbruck/Listen/HEB_Liste_Schnell.xlsx"]
    for file in files:
        if sys.platform == 'linux':
            subprocess.call(["xdg-open",file])
        else:
            os.system("start " + file)

def store_reasons_dict(reasons_pcr_dict, reasons_schnell_dict):
    with open('./modules/testgruende.json', 'w') as f:
        json.dump(reasons_pcr_dict, f)
    with open('./modules/testgruende_schnell.json', 'w') as d:
        json.dump(reasons_schnell_dict, d)

def load_reasons_dict():
    with open('./modules/testgruende.json') as f:
        reasons_pcr_dict = json.load(f)
    with open('./modules/testgruende_schnell.json') as d:
        reasons_schnell_dict = json.load(d)
    return reasons_schnell_dict, reasons_pcr_dict
    
def replace_all(html_text, dic):
    """
    Tauscht im HTML Template Keywords mit den echten Daten des Dictionaries
    """
    for i, j in dic.items():
        html_text = html_text.replace(str(i), str(j))
    return html_text


def define_testgrund(dict_reasons, testgrund):
    """
    Returned den Testgrund Paraghraphen zum Stichwort im Girona Export

    reasons_dict = dict mit Stichwort als key und paragraphen als value
    testgrund = Tesgrund Stichwort der Girona Excel
    
    """
    try:
        reason = dict_reasons[testgrund]
    except KeyError:
        reason = ""
    return reason

def remove_columns(df):
    """Remove all columns which aren't needed for the list"""
    needed_cols = ["Termine", "GF_Familienname"	,"GF_Vorname","GF_Geburtsdatum","GF_Strasse","GF_Hausnr","GF_PLZ", "GF_Ort", "GF_Email", "Testgrund", "Testgrund_Schnelltest"]

    for column in df.columns:
        if column not in needed_cols:
            del df[column]
            
    return df

def add_last_two_cols(df):
    """Adds a number and Befund column to the excel List"""

    new_column = [i + 1 for i in range(len(df))]
    df.insert(loc = len(df.columns),
              column='Nr.',
              value=new_column)
    df.insert(loc=len(df.columns),
              column='  Befund  ',
              value=None)

    return df


def filter_df_by_time(df, starttime=1700, endtime=1900):
    starttime = int(starttime)
    endtime = int(endtime)
    for i in range(len(df)):
        if (df.loc[i, "Termine_copy"]) < starttime or (df.loc[i, "Termine_copy"]) > endtime:
            df.drop(index=i, inplace=True)
    return df

def cutoff_two_appointments(df):
    # cut off if there are multiple appointments on one person and sort by the first one
    df['Termine'] = [x[:16] for x in df['Termine']]
    return df

def sort_df_ascending(df):
    """
    Sortiert dataframe nach Uhrzeit. Es kopiert die Spalte, damit die Originaldatumsformatierung im Ausdruck passt.
    Termine_copy enthält nach der Sortierung nur die Uhrzeit als int
    """
    #sort dataframe by test time in ascending order
    df['Termine_copy'] = pd.to_datetime(df['Termine'], format="%d.%m.%Y %H:%M", errors='coerce')
    df = df.sort_values(by='Termine_copy',ascending=True)
    df["Termine_copy"] =  df["Termine_copy"].dt.strftime("%H%M").apply(int)
    return df

def replace_headers(df):
    header_names = {
    "Termine": "Termine",
    "GF_Familienname": "Familienname",
    "GF_Vorname": "Vorname",
    "GF_Geburtsdatum":"Geburtsdatum",
    "GF_Strasse": "Strasse",
    "GF_Hausnr": "HN",
    "GF_Email": "Email",
    "GF_PLZ": "PLZ",
    "GF_Ort": "Ort"
    }

    df.rename(columns=header_names, inplace=True)
    return df

def print_to_excel(df, name, pcr):
    """
    prints df to excel with appropriate name

    df: dataframe to be printed
    name: name to be used in file name
    """
    if pcr:
        del df["Testgrund_Schnelltest"]
    else:
        del df["Testgrund"]
        
    df.index = list(range(1, len(df) + 1))
    with pd.ExcelWriter(r''+ name+'.xlsx') as writer:
        df.to_excel(writer)
    
    ## Styling wie Borders und duplicate markieren
    __style_excel(df, name + ".xlsx")
    
def define_ressource(ressource):
    if ressource == "Hersbruck Schnelltest" or ressource == "Hersbruck PCR-Test":
            ressource = "Testzentrum Hersbruck"

    elif ressource == "Altdorf Schnelltest" or ressource == "Altdorf PCR-Test":
        ressource = "Testzentrum Altdorf"

    elif ressource == "Lauf Schnelltest" or ressource == "Lauf PCR-Test":
        ressource = "Testzentrum Lauf"
    return ressource

def split_df_teststellen(df):
    """
    Creates the Dataframe splits the big Export into the needed dataframes and returns them as tuple
    """
    #make dataframe for all facilities
    df_altdorf = pd.DataFrame()
    df_hersbruck = pd.DataFrame()
    df_lauf = pd.DataFrame()
    df_altdorf_pcr = pd.DataFrame()
    df_hersbruck_pcr = pd.DataFrame()
    df_lauf_pcr = pd.DataFrame()

    df_hersbruck = df[df.Standort == "91217 Hersbruck"]
    df_hersbruck_pcr = df_hersbruck[df_hersbruck.Testart == "PCR-Test"]
    df_hersbruck_pcr.reset_index(inplace=True, drop=True)
    df_hersbruck = df_hersbruck[df_hersbruck.Testart != "PCR-Test"]
    df_hersbruck.reset_index(inplace=True, drop=True)

    df_altdorf = df[df.Standort == "90518  Altdorf"]
    df_altdorf_pcr = df_altdorf[df_altdorf.Testart == "PCR-Test"]
    df_altdorf_pcr.reset_index(inplace=True, drop=True)
    df_altdorf = df_altdorf[df_altdorf.Testart != "PCR-Test"]
    df_altdorf.reset_index(inplace=True, drop=True)

    df_lauf = df[df.Standort == "91207 Lauf"]
    df_lauf_pcr = df_lauf[df_lauf.Testart == "PCR-Test"]
    df_lauf_pcr.reset_index(inplace=True, drop=True)
    df_lauf = df_lauf[df_lauf.Testart != "PCR-Test"]
    df_lauf.reset_index(inplace=True, drop=True)

    return (df_hersbruck, df_hersbruck_pcr, df_altdorf, df_altdorf_pcr, df_lauf, df_lauf_pcr)

def generate_html(df, name, dict_reasons, pcr=True):
    """
    Generates the HTML-File for a Girona csv Dataframe 

    df: dataframe to be converted to HTML
    name: name to be used in HTML filename
    dict_reasons: dictionary of test reasons witht their paragraphs
    """
    dict={}

    html_text = html_head

    if pcr:
        spalte = "Testgrund"
    else:
        spalte = "Testgrund_Schnelltest"


    if len(df) > 0:
        # Die Dataframes werden davor gesplittet also muss die Ressource überall gleich sein deswegen nehmen wir einfach die Ressource bei Index 0
        html_text += """<h1 class="description-title" style="text-align:center;" > """+ df.loc[0, 'Ressource'] +  """</h1> """

    for i in range(len(df)):

        dict['{{Termine}}'] = df.loc[i,'Termine']
        dict['{{GF_Vorname}}'] = df.loc[i,'GF_Vorname']
        dict['{{GF_Familienname}}'] = df.loc[i, 'GF_Familienname']
        dict['{{GF_Geburtsdatum}}'] = df.loc[i, 'GF_Geburtsdatum']
        dict['{{GF_Strasse}}'] = df.loc[i, 'GF_Strasse']
        dict['{{GF_Hausnr}}'] =  df.loc[i, 'GF_Hausnr']
        dict['{{GF_PLZ}}'] =  df.loc[i, 'GF_PLZ']
        dict['{{GF_Ort}}'] = df.loc[i, 'GF_Ort']
        dict['{{GF_Telefon}}'] = df.loc[i, 'GF_Telefon']
        dict['{{GF_Email}}'] = df.loc[i, 'GF_Email']
        dict['{{Testart}}'] = df.loc[i, 'Testart']
        dict['{{Ressource}}'] = define_ressource(df.loc[i, 'Ressource'])
        dict['{{Grund}}'] = define_testgrund(dict_reasons ,df.loc[i, spalte])
        dict['{{Testnummer}}'] = i + 1

        if dict['{{Testart}}'] == "Schnelltest":
            dict['{{ID}}'] = 'Test-ID:AT001/20 BfArM'
        else:
            dict['{{ID}}'] = ''

        if pcr:
            html_for_person = html_page_template
        else:
            html_for_person = html_page_template + html_selbstauskunft

        html_text += replace_all(html_for_person, dict)
    
    #wenn das Dataframe leer ist, dann sind hier keine Tests angemeldet
    if len(df) == 0:
        html_text +=""" <div style="border: 5px solid red; width: 50%;"><h1 style="text-align: center;"><strong><em>Heute wurden hier keine Tests angemeldet!</em></strong></h1></div>"""

    html_text += html_tail
    #open text file
    text_file = open("" + name + ".html", "w", encoding="utf-8")
    #write string to file
    text_file.write(html_text)
    #close file
    text_file.close()

def duplicates_of_persons(df):
    """
    Counts the amount of times a person with the same name is in the Dataframe and returns a list of people where the amount is bigger than 1
    """
    name_series = df['GF_Vorname'].apply(str.lower) + " " + df['GF_Familienname'].apply(str.lower) + " " + df["Ressource"]
    counted_persons = name_series.value_counts()
    counted_persons = counted_persons.where(counted_persons > 1)
    counted_persons.dropna(inplace=True)
    return counted_persons

def find_testgruende(df):
    """
    Finds all test reasons inside the dataframe 

    There are two relevant columns for this one for pcr and one for lateral flow
    """
    pcr_reasons = df["Testgrund"].value_counts()
    schnell_reasons = df["Testgrund_Schnelltest"].value_counts()
    return schnell_reasons, pcr_reasons